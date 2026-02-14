#!/usr/bin/env python3
"""
Rebuild KCET cutoff dataset from strict sources:
- 2023 and 2024 from XLSX files
- 2025 directly from official PDF files

This script keeps course names exactly as present in source files
(only whitespace-normalized), deduplicates by full cutoff identity,
and writes the rebuilt dataset to all runtime JSON locations.
"""

from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl
import pdfplumber


ROOT = Path(__file__).resolve().parent.parent
PUBLIC_DIR = ROOT / "public"

XLSX_FILES = sorted(PUBLIC_DIR.glob("kcet-202[34]*-cutoffs.xlsx"))
PDF_2025_FILES = sorted(PUBLIC_DIR.glob("cutoffs/kcet-2025*.pdf"))

TARGET_JSON_PATHS = [
    PUBLIC_DIR / "data" / "kcet_cutoffs_master.json",
    PUBLIC_DIR / "data" / "kcet_cutoffs_high_volume.json",
    PUBLIC_DIR / "data" / "kcet_cutoffs_consolidated.json",
    PUBLIC_DIR / "kcet_cutoffs_master.json",
    PUBLIC_DIR / "kcet_cutoffs_high_volume.json",
    PUBLIC_DIR / "kcet_cutoffs_consolidated.json",
    PUBLIC_DIR / "kcet_cutoffs.json",
]

CATEGORY_CODES = {
    "1G",
    "1K",
    "1R",
    "2AG",
    "2AK",
    "2AR",
    "2BG",
    "2BK",
    "2BR",
    "3AG",
    "3AK",
    "3AR",
    "3BG",
    "3BK",
    "3BR",
    "GM",
    "GMK",
    "GMR",
    "SCG",
    "SCK",
    "SCR",
    "STG",
    "STK",
    "STR",
    "GMP",
    "NRI",
    "OPN",
    "OTH",
}

_RE_SPACES = re.compile(r"\s+")
_RE_ECODE = re.compile(r"\b(E\d{3})\b", re.IGNORECASE)
_RE_NUMERIC = re.compile(r"^[\d.,\-\s]+$")


def raw_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").strip()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ")
    text = text.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    text = _RE_SPACES.sub(" ", text).strip()
    return text


def parse_year_round_from_filename(filename: str) -> Tuple[Optional[str], Optional[str]]:
    lower = filename.lower()
    year_match = re.search(r"(2023|2024)", lower)
    if not year_match:
        return None, None
    year = year_match.group(1)
    if "mock" in lower:
        return year, "MOCK"
    if "round3" in lower or "extended" in lower:
        return year, "R3"
    if "round2" in lower:
        return year, "R2"
    if "round1" in lower:
        return year, "R1"
    return year, "R1"


def parse_round_2025_from_filename(filename: str) -> Optional[str]:
    lower = clean_text(filename).lower()
    if not lower:
        return None
    if "mock" in lower:
        return "MOCK"
    if "round 3" in lower or "round3" in lower or "extended" in lower:
        return "R3"
    if "round 2" in lower or "round2" in lower:
        return "R2"
    if "round 1" in lower or "round1" in lower:
        return "R1"
    return None


def parse_rank(value: object) -> Optional[int]:
    text = clean_text(value)
    if text in {"", "-", "--", "NA", "N/A", "None", "nan"}:
        return None
    text = text.replace(",", "")

    # OCR/table extraction often introduces spaces in decimals (e.g. "142544. 5")
    # or digit chunks (e.g. "34096.87 5"). Normalize those first.
    text = re.sub(r"\s*\.\s*", ".", text)
    text = re.sub(r"(?<=\d)\s+(?=\d)", "", text)

    # If the whole string still isn't a clean number, take the first numeric token.
    number_match = re.search(r"\d+(?:\.\d+)?", text)
    if not number_match:
        return None

    try:
        rank = int(float(number_match.group(0)))
    except ValueError:
        return None
    if rank < 1 or rank > 500000:
        return None
    return rank


def extract_college_from_row(
    row_values: List[object], known_college_names: Dict[str, str]
) -> Optional[Tuple[str, str]]:
    for idx, raw_value in enumerate(row_values):
        text = clean_text(raw_value)
        if not text:
            continue
        match = _RE_ECODE.search(text)
        if not match:
            continue

        code = match.group(1).upper()
        trailing = clean_text(text[match.end() :]).strip(" -:.,;()[]")
        trailing = re.sub(r"^(college\s*:?)\s*", "", trailing, flags=re.IGNORECASE)

        if trailing:
            known_college_names[code] = trailing
            return code, trailing

        # Name not in same cell. Try nearby cells in same row.
        for next_idx in range(max(0, idx - 2), min(len(row_values), idx + 8)):
            if next_idx == idx:
                continue
            candidate = clean_text(row_values[next_idx])
            if not candidate:
                continue
            if _RE_ECODE.search(candidate):
                continue
            candidate_upper = candidate.upper()
            if candidate_upper in CATEGORY_CODES or _RE_NUMERIC.match(candidate):
                continue
            if len(candidate) >= 4:
                known_college_names[code] = candidate
                return code, candidate

        # Keep previously known name for this code if available.
        if code in known_college_names:
            return code, known_college_names[code]
        return code, f"College {code}"

    return None


def extract_category_map(row_values: List[object]) -> Dict[int, str]:
    category_map: Dict[int, str] = {}
    for idx, raw_value in enumerate(row_values):
        value = clean_text(raw_value).upper()
        if value in CATEGORY_CODES:
            category_map[idx] = value
    if len(category_map) >= 5:
        return category_map
    return {}


def looks_like_non_course(text: str) -> bool:
    upper = text.upper()
    if upper in CATEGORY_CODES:
        return True
    if _RE_ECODE.search(upper):
        return True
    if _RE_NUMERIC.match(text):
        return True
    if text in {"-", "--"}:
        return True
    blocked = (
        "ENGINEERING CUTOFF",
        "ALLOTMENT",
        "COURSE NAME",
        "COURSE",
        "BRANCH NAME",
        "BRANCH",
    )
    return any(token in upper for token in blocked)


def extract_course_name(row_values: List[object], category_cols: Iterable[int]) -> Optional[str]:
    category_cols = sorted(category_cols)
    if not category_cols:
        return None
    first_cat_col = category_cols[0]

    # Course names are consistently on the left side.
    for idx in range(0, min(first_cat_col, 8)):
        original = raw_text(row_values[idx] if idx < len(row_values) else "")
        text = clean_text(original)
        if not original or not text or looks_like_non_course(text):
            continue
        if re.search(r"[A-Za-z]", text):
            return original

    # Fallback: scan a bit wider before categories.
    for idx in range(0, min(first_cat_col + 1, 14)):
        original = raw_text(row_values[idx] if idx < len(row_values) else "")
        text = clean_text(original)
        if not original or not text or looks_like_non_course(text):
            continue
        if re.search(r"[A-Za-z]", text):
            return original

    return None


def extract_2023_2024_from_xlsx() -> Tuple[List[dict], dict]:
    results: List[dict] = []
    source_stats: Dict[str, dict] = {}
    known_college_names: Dict[str, str] = {}
    all_detected_codes: set[str] = set()

    for xlsx_path in XLSX_FILES:
        year, round_value = parse_year_round_from_filename(xlsx_path.name)
        if not year or not round_value:
            continue

        workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        file_count = 0
        file_codes = set()
        current_code: Optional[str] = None
        current_name: Optional[str] = None
        current_category_map: Dict[int, str] = {}

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                row_values = list(row)
                college_data = extract_college_from_row(row_values, known_college_names)
                if college_data:
                    current_code, current_name = college_data
                    file_codes.add(current_code)
                    all_detected_codes.add(current_code)
                    current_category_map = {}
                    continue

                category_map = extract_category_map(row_values)
                if category_map and current_code:
                    current_category_map = category_map
                    continue

                if not current_code or not current_name or not current_category_map:
                    continue

                course_name = extract_course_name(row_values, current_category_map.keys())
                if not course_name:
                    continue

                for col_idx, category in current_category_map.items():
                    if col_idx >= len(row_values):
                        continue
                    rank = parse_rank(row_values[col_idx])
                    if rank is None:
                        continue

                    results.append(
                        {
                            "institute": current_name,
                            "institute_code": current_code,
                            "course": course_name,
                            "category": category,
                            "cutoff_rank": rank,
                            "year": year,
                            "round": round_value,
                        }
                    )
                    file_count += 1

        workbook.close()
        source_stats[xlsx_path.name] = {
            "entries": file_count,
            "detected_institutes": len(file_codes),
        }

    return results, {"files": source_stats, "all_detected_institutes": len(all_detected_codes)}


def extract_college_headers_from_pdf_page(page: pdfplumber.page.Page) -> List[dict]:
    words = page.extract_words(keep_blank_chars=True) or []
    lines: Dict[int, List[str]] = {}
    for word in words:
        top = round(float(word.get("top", 0)))
        text = word.get("text", "")
        lines.setdefault(top, []).append(text)

    headers: List[dict] = []
    for top in sorted(lines.keys()):
        line_text = clean_text(" ".join(lines[top]))
        match = re.search(
            r"College\s*:\s*[\(\[]?([A-Z]\d{3})[\)\]]?\s*(.*)",
            line_text,
            re.IGNORECASE,
        )
        if not match:
            continue

        code = match.group(1).upper()
        name = clean_text(match.group(2))
        if "Course Name" in name:
            name = clean_text(name.split("Course Name")[0])
        name = name.strip(" -:;,.")

        if code:
            headers.append(
                {
                    "top": top,
                    "code": code,
                    "name": name if name else f"College {code}",
                }
            )
    return headers


def extract_2025_from_pdfs() -> Tuple[List[dict], dict]:
    if not PDF_2025_FILES:
        raise FileNotFoundError("No 2025 cutoff PDF files found in public/cutoffs/")

    results: List[dict] = []
    file_stats: Dict[str, dict] = {}
    all_codes = set()

    for pdf_path in PDF_2025_FILES:
        round_value = parse_round_2025_from_filename(pdf_path.name)
        if not round_value:
            continue

        file_count = 0
        file_codes = set()
        current_code: Optional[str] = None
        current_name: Optional[str] = None

        with pdfplumber.open(pdf_path) as pdf:
            page_count = len(pdf.pages)

            for page in pdf.pages:
                headers = extract_college_headers_from_pdf_page(page)
                # Carry forward the last visible college header on this page
                # even if this page has no table (common in some 2025 PDFs).
                if headers:
                    last_header = max(headers, key=lambda h: h["top"])
                    current_code = last_header["code"]
                    current_name = last_header["name"]

                tables = page.find_tables() or []

                for table in tables:
                    table_top = table.bbox[1] if table.bbox else 0
                    candidates = [h for h in headers if h["top"] < table_top]
                    if candidates:
                        nearest = max(candidates, key=lambda h: h["top"])
                        current_code = nearest["code"]
                        current_name = nearest["name"]

                    if not current_code or not current_name:
                        continue

                    table_rows = table.extract() or []
                    if not table_rows:
                        continue

                    category_map: Dict[int, str] = {}
                    header_row_idx = -1
                    for row_idx, row_values in enumerate(table_rows[:12]):
                        row_list = list(row_values) if row_values else []
                        detected = extract_category_map(row_list)
                        if len(detected) >= 8:
                            category_map = detected
                            header_row_idx = row_idx
                            break

                    if header_row_idx < 0 or not category_map:
                        continue

                    for row_values in table_rows[header_row_idx + 1 :]:
                        row_list = list(row_values) if row_values else []
                        course_original = raw_text(row_list[0] if row_list else "")
                        course_check = clean_text(course_original)
                        if not course_original or not course_check or looks_like_non_course(course_check):
                            continue

                        for col_idx, category in category_map.items():
                            if col_idx >= len(row_list):
                                continue
                            rank = parse_rank(row_list[col_idx])
                            if rank is None:
                                continue

                            results.append(
                                {
                                    "institute": current_name,
                                    "institute_code": current_code,
                                    "course": course_original,
                                    "category": category,
                                    "cutoff_rank": rank,
                                    "year": "2025",
                                    "round": round_value,
                                }
                            )
                            file_count += 1
                            file_codes.add(current_code)
                            all_codes.add(current_code)

        file_stats[pdf_path.name] = {
            "entries": file_count,
            "detected_institutes": len(file_codes),
            "pages": page_count,
        }

    return results, {"files": file_stats, "all_detected_institutes": len(all_codes)}


def dedupe_rows(rows: List[dict]) -> Tuple[List[dict], int]:
    seen = set()
    unique_rows: List[dict] = []
    duplicates = 0

    for row in rows:
        key = (
            row["institute_code"],
            row["course"],
            row["category"],
            row["year"],
            row["round"],
            row["cutoff_rank"],
        )
        if key in seen:
            duplicates += 1
            continue
        seen.add(key)
        unique_rows.append(row)

    unique_rows.sort(
        key=lambda r: (
            r["year"],
            r["round"],
            r["institute_code"],
            r["course"],
            r["category"],
            r["cutoff_rank"],
        )
    )
    return unique_rows, duplicates


def build_metadata(rows: List[dict], duplicates_removed: int, source_info: dict) -> dict:
    by_year = Counter(r["year"] for r in rows)
    by_round = Counter(r["round"] for r in rows)
    by_year_round = Counter((r["year"], r["round"]) for r in rows)
    institutes = sorted({r["institute_code"] for r in rows})
    categories = sorted({r["category"] for r in rows})
    courses = sorted({r["course"] for r in rows})

    records_by_year_round = {
        f"{year}_{round_value}": count
        for (year, round_value), count in sorted(by_year_round.items())
    }

    return {
        "last_updated": __import__("datetime").datetime.now().isoformat(),
        "source_type": "strict_rebuild_xlsx_2023_2024_plus_pdf_2025",
        "total_entries": len(rows),
        "duplicates_removed": duplicates_removed,
        "total_institutes": len(institutes),
        "total_courses": len(courses),
        "total_categories": len(categories),
        "years_covered": sorted(by_year.keys()),
        "rounds_covered": sorted(by_round.keys()),
        "records_by_year": {k: by_year[k] for k in sorted(by_year.keys())},
        "records_by_round": {k: by_round[k] for k in sorted(by_round.keys())},
        "records_by_year_round": records_by_year_round,
        "dedupe_key": "institute_code+course+category+year+round+cutoff_rank",
        "sources": source_info,
    }


def write_output(payload: dict) -> None:
    for path in TARGET_JSON_PATHS:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False)


def main() -> None:
    if not XLSX_FILES:
        raise FileNotFoundError("No 2023/2024 XLSX files found in public/")

    xlsx_rows, xlsx_stats = extract_2023_2024_from_xlsx()
    rows_2025, pdf_stats = extract_2025_from_pdfs()

    merged_rows = xlsx_rows + rows_2025
    unique_rows, duplicates_removed = dedupe_rows(merged_rows)

    metadata = build_metadata(
        unique_rows,
        duplicates_removed,
        {
            "xlsx_2023_2024": xlsx_stats,
            "pdf_2025": pdf_stats,
            "xlsx_files": [p.name for p in XLSX_FILES],
            "pdf_2025_files": [str(p.relative_to(ROOT)) for p in PDF_2025_FILES],
        },
    )

    payload = {"metadata": metadata, "cutoffs": unique_rows}
    write_output(payload)

    print("Rebuild complete")
    print(f"  total entries: {metadata['total_entries']:,}")
    print(f"  duplicates removed: {duplicates_removed:,}")
    print(f"  records by year: {metadata['records_by_year']}")
    print(f"  records by round: {metadata['records_by_round']}")
    print("  updated files:")
    for path in TARGET_JSON_PATHS:
        print(f"    - {path.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
